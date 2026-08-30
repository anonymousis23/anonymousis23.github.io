#!/usr/bin/env python3
"""Export one participant-level administrative audit workbook per study."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_DATABASE = Path(__file__).resolve().parent / "responses.db"
DEFAULT_OUTPUT = Path("/data/waris/data/PHONOS_TASLP26_admin_audit_20260828")
STUDIES = {
    "phonos_taslp26_mos": {
        "filename": "PHONOS_TASLP26_MOS_participant_audit.xlsx",
        "title": "PHONOS TASLP26 MOS",
        "audit_task_type": "MOS (1-5)",
        "response_fields": (
            ("mos_rating", "MOS Rating"),
            ("mos_label", "MOS Label"),
            ("distortion_label", "Distortion Label"),
            ("response_ts_ms", "Response Timestamp (UTC)"),
        ),
    },
    "phonos_taslp26_accent_british": {
        "filename": "PHONOS_TASLP26_Accent_British_participant_audit.xlsx",
        "title": "PHONOS TASLP26 American-British Accent Verification",
        "audit_task_type": "Accent verification (American/British)",
        "response_fields": (
            ("accent_choice", "Accent Choice"),
            ("confidence", "Confidence (1-7)"),
            ("response_ts_ms", "Response Timestamp (UTC)"),
        ),
    },
    "phonos_taslp26_accent_indian": {
        "filename": "PHONOS_TASLP26_Accent_Indian_participant_audit.xlsx",
        "title": "PHONOS TASLP26 American-Indian Accent Verification",
        "audit_task_type": "Accent verification (American/Indian)",
        "response_fields": (
            ("accent_choice", "Accent Choice"),
            ("confidence", "Confidence (1-7)"),
            ("response_ts_ms", "Response Timestamp (UTC)"),
        ),
    },
    "phonos_taslp26_accent_spanish": {
        "filename": "PHONOS_TASLP26_Accent_Spanish_participant_audit.xlsx",
        "title": "PHONOS TASLP26 American-Spanish Accent Verification",
        "audit_task_type": "Accent verification (American/Spanish-accented English)",
        "response_fields": (
            ("accent_choice", "Accent Choice"),
            ("confidence", "Confidence (1-7)"),
            ("response_ts_ms", "Response Timestamp (UTC)"),
        ),
    },
}

METADATA_HEADERS = [
    "Participant Number",
    "PROLIFIC_PID",
    "Prolific STUDY_ID",
    "Prolific SESSION_ID",
    "Effective Participant ID",
    "Survey Participant ID",
    "Submission ID",
    "Internal Study ID",
    "Stored Task Type",
    "Audit Task Type",
    "Target Accent",
    "Received At (UTC)",
    "Started At (UTC)",
    "Submitted At (UTC)",
    "Study Duration (minutes)",
    "Response Count",
    "Expected Response Count",
    "Completion Rate (%)",
    "Randomized Order Seed",
    "Page Size",
    "Cooldown Seconds",
    "Audio Setup",
    "Participant Comments",
    "Page URL",
    "User Agent",
    "Raw Payload SHA256",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--database", type=Path, default=DEFAULT_DATABASE)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def parse_payload(value: Any) -> dict[str, Any]:
    if isinstance(value, str):
        return json.loads(value)
    return value or {}


def iso_from_ms(value: Any) -> str:
    if value in (None, ""):
        return ""
    return datetime.fromtimestamp(float(value) / 1000.0, tz=timezone.utc).isoformat()


def received_iso(value: Any) -> str:
    if value in (None, ""):
        return ""
    text = str(value)
    parsed = datetime.fromisoformat(text)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc).isoformat()


def duration_minutes(started: Any, submitted: Any) -> float | str:
    if started in (None, "") or submitted in (None, ""):
        return ""
    return round((float(submitted) - float(started)) / 60000.0, 3)


def is_explicit_test(payload: dict[str, Any]) -> tuple[bool, str]:
    survey = payload.get("post_survey") or {}
    survey_id = str(survey.get("participant_id") or "").strip()
    comments = str(survey.get("comments") or "").strip()
    reasons = []
    if survey_id.casefold() == "test":
        reasons.append("survey participant ID is TEST")
    if re.search(r"\btest\b", comments, flags=re.IGNORECASE):
        reasons.append("participant comments explicitly mark TEST/IGNORE")
    return bool(reasons), "; ".join(reasons)


def load_database(database: Path) -> tuple[list[dict], list[dict]]:
    connection = sqlite3.connect(f"file:{database.resolve()}?mode=ro", uri=True)
    connection.row_factory = sqlite3.Row
    submissions = [
        dict(row)
        for row in connection.execute("SELECT * FROM submissions ORDER BY received_at, id")
    ]
    trials = [
        dict(row)
        for row in connection.execute(
            "SELECT * FROM trial_responses ORDER BY submission_id, display_index, qid"
        )
    ]
    connection.close()
    return submissions, trials


def canonical_json(payload: dict[str, Any]) -> str:
    return json.dumps(payload, sort_keys=True, separators=(",", ":"), ensure_ascii=True)


def build_question_metadata(trials: list[dict]) -> list[dict]:
    fields = [
        "qid",
        "display_index",
        "page",
        "condition",
        "condition_label",
        "expected_accent",
        "audio",
        "source_index",
    ]
    metadata: dict[str, dict] = {}
    for trial in trials:
        candidate = {field: trial.get(field) for field in fields}
        qid = str(trial.get("qid") or "")
        if qid in metadata and candidate != metadata[qid]:
            raise RuntimeError(f"Question metadata conflict for {qid}")
        metadata[qid] = candidate
    return sorted(metadata.values(), key=lambda row: (row["display_index"], row["qid"]))


def build_participant_row(
    number: int,
    submission: dict,
    payload: dict,
    trials: list[dict],
    study: dict,
    question_ids: list[str],
) -> dict[str, Any]:
    participant = payload.get("participant") or {}
    survey = payload.get("post_survey") or {}
    started = submission.get("started_at_ms")
    submitted = submission.get("submitted_at_ms")
    row: dict[str, Any] = {
        "Participant Number": number,
        "PROLIFIC_PID": submission.get("prolific_pid") or participant.get("PROLIFIC_PID") or "",
        "Prolific STUDY_ID": submission.get("prolific_study_id") or participant.get("STUDY_ID") or "",
        "Prolific SESSION_ID": submission.get("prolific_session_id") or participant.get("SESSION_ID") or "",
        "Effective Participant ID": submission.get("participant_id") or "",
        "Survey Participant ID": survey.get("participant_id") or "",
        "Submission ID": submission.get("id") or "",
        "Internal Study ID": submission.get("study_id") or "",
        "Stored Task Type": submission.get("task_type") or "",
        "Audit Task Type": study["audit_task_type"],
        "Target Accent": submission.get("target_accent") or "",
        "Received At (UTC)": received_iso(submission.get("received_at")),
        "Started At (UTC)": iso_from_ms(started),
        "Submitted At (UTC)": iso_from_ms(submitted),
        "Study Duration (minutes)": duration_minutes(started, submitted),
        "Response Count": len(trials),
        "Expected Response Count": len(question_ids),
        "Completion Rate (%)": round(100.0 * len(trials) / len(question_ids), 3),
        "Randomized Order Seed": payload.get("randomized_order_seed"),
        "Page Size": payload.get("page_size"),
        "Cooldown Seconds": payload.get("cooldown_seconds"),
        "Audio Setup": survey.get("audio_setup") or "",
        "Participant Comments": survey.get("comments") or "",
        "Page URL": submission.get("page_url") or payload.get("page_url") or "",
        "User Agent": submission.get("user_agent") or payload.get("user_agent") or "",
        "Raw Payload SHA256": hashlib.sha256(canonical_json(payload).encode()).hexdigest(),
    }
    by_qid = {str(trial.get("qid") or ""): trial for trial in trials}
    for qid in question_ids:
        trial = by_qid.get(qid, {})
        for field, label in study["response_fields"]:
            value = trial.get(field, "")
            if field == "response_ts_ms":
                value = iso_from_ms(value)
            row[f"{qid} {label}"] = "" if value is None else value
    return row


def set_header_style(sheet, metadata_columns: int, excluded: bool = False) -> None:
    metadata_fill = PatternFill("solid", fgColor="7A2525" if excluded else "17324D")
    response_fill = PatternFill("solid", fgColor="176B3A")
    for index, cell in enumerate(sheet[1], start=1):
        cell.fill = metadata_fill if index <= metadata_columns else response_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 48


def populate_wide_sheet(
    sheet,
    rows: list[dict[str, Any]],
    response_headers: list[str],
    excluded: bool = False,
) -> None:
    prefix = ["Exclusion Reason"] if excluded else []
    headers = prefix + METADATA_HEADERS + response_headers
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    set_header_style(sheet, len(prefix) + len(METADATA_HEADERS), excluded=excluded)
    sheet.freeze_panes = f"{get_column_letter(len(prefix) + len(METADATA_HEADERS) + 1)}2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, sheet.max_row)}"
    for index, header in enumerate(headers, start=1):
        if header in {"Participant Comments", "Page URL", "User Agent"}:
            width = 42
        elif header.endswith("Timestamp (UTC)") or " At (UTC)" in header:
            width = 27
        elif index > len(prefix) + len(METADATA_HEADERS):
            width = 20
        else:
            width = min(max(len(header) + 2, 13), 28)
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def populate_question_sheet(sheet, metadata: list[dict]) -> None:
    headers = [
        "QID",
        "Display Index",
        "Page",
        "Condition",
        "Condition Label",
        "Expected Accent",
        "Audio Path",
        "Source Index",
    ]
    sheet.append(headers)
    for item in metadata:
        sheet.append(
            [
                item.get("qid"),
                item.get("display_index"),
                item.get("page"),
                item.get("condition"),
                item.get("condition_label"),
                item.get("expected_accent"),
                item.get("audio"),
                item.get("source_index"),
            ]
        )
    set_header_style(sheet, len(headers))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:H{sheet.max_row}"
    for index, width in enumerate([12, 15, 10, 24, 52, 28, 58, 14], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def populate_dictionary_sheet(
    sheet,
    study_id: str,
    study: dict,
    included_count: int,
    excluded_count: int,
    database: Path,
) -> None:
    generated = datetime.now(timezone.utc).isoformat()
    entries = [
        ("Workbook title", study["title"]),
        ("Internal study ID", study_id),
        ("Included participant rows", included_count),
        ("Excluded explicit test rows", excluded_count),
        ("Expected responses per participant", 60),
        ("Generated at (UTC)", generated),
        ("Source database", str(database.resolve())),
        ("Inclusion rule", "Complete 60-response submissions not explicitly marked TEST/IGNORE."),
        ("Participants sheet", "One row per included participant, followed by all item responses in QID order."),
        ("Question Metadata sheet", "Maps each QID to its condition, expected accent, and audio path."),
        ("Excluded Tests sheet", "Explicit test submissions retained for audit transparency but excluded from the cohort."),
        ("PROLIFIC_PID", "Participant identifier supplied in the Prolific launch URL."),
        ("Prolific STUDY_ID", "Prolific platform study identifier supplied in the launch URL."),
        ("Prolific SESSION_ID", "Prolific platform session identifier supplied in the launch URL."),
        ("Internal Study ID", "PHONOS application study ID; distinct from Prolific STUDY_ID."),
        ("Response Timestamp", "UTC conversion of the browser response timestamp stored in milliseconds."),
        ("Raw Payload SHA256", "SHA-256 of canonicalized stored JSON for traceability."),
    ]
    sheet.append(["Field", "Description / Value"])
    for entry in entries:
        sheet.append(entry)
    set_header_style(sheet, 2)
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 105
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_workbook(
    path: Path,
    study_id: str,
    study: dict,
    included: list[dict],
    excluded: list[dict],
    question_metadata: list[dict],
    database: Path,
) -> None:
    question_ids = [item["qid"] for item in question_metadata]
    response_headers = [
        f"{qid} {label}"
        for qid in question_ids
        for _field, label in study["response_fields"]
    ]
    workbook = Workbook()
    participants_sheet = workbook.active
    participants_sheet.title = "Participants"
    populate_wide_sheet(participants_sheet, included, response_headers)
    question_sheet = workbook.create_sheet("Question Metadata")
    populate_question_sheet(question_sheet, question_metadata)
    excluded_sheet = workbook.create_sheet("Excluded Tests")
    populate_wide_sheet(excluded_sheet, excluded, response_headers, excluded=True)
    dictionary_sheet = workbook.create_sheet("Data Dictionary")
    populate_dictionary_sheet(
        dictionary_sheet,
        study_id,
        study,
        len(included),
        len(excluded),
        database,
    )
    workbook.save(path)


def main() -> None:
    args = parse_args()
    args.output_root.mkdir(parents=True, exist_ok=True)
    submissions, trials = load_database(args.database)
    trials_by_submission: dict[str, list[dict]] = {}
    for trial in trials:
        trials_by_submission.setdefault(trial["submission_id"], []).append(trial)

    export_summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_database": str(args.database.resolve()),
        "output_root": str(args.output_root.resolve()),
        "studies": {},
    }
    for study_id, study in STUDIES.items():
        study_submissions = [row for row in submissions if row["study_id"] == study_id]
        study_trials = [trial for row in study_submissions for trial in trials_by_submission.get(row["id"], [])]
        question_metadata = build_question_metadata(study_trials)
        question_ids = [row["qid"] for row in question_metadata]
        if len(question_ids) != 60:
            raise RuntimeError(f"{study_id}: expected 60 QIDs, found {len(question_ids)}")

        included_rows = []
        excluded_rows = []
        for submission in study_submissions:
            payload = parse_payload(submission["raw_payload"])
            submission_trials = trials_by_submission.get(submission["id"], [])
            if len(submission_trials) != 60:
                raise RuntimeError(f"{submission['id']}: expected 60 responses, found {len(submission_trials)}")
            is_test, reason = is_explicit_test(payload)
            destination = excluded_rows if is_test else included_rows
            participant_number = len(destination) + 1
            row = build_participant_row(
                participant_number,
                submission,
                payload,
                submission_trials,
                study,
                question_ids,
            )
            if is_test:
                row["Exclusion Reason"] = reason
            destination.append(row)

        filename = study["filename"]
        workbook_path = args.output_root / filename
        write_workbook(
            workbook_path,
            study_id,
            study,
            included_rows,
            excluded_rows,
            question_metadata,
            args.database,
        )
        export_summary["studies"][study_id] = {
            "workbook": filename,
            "included_participants": len(included_rows),
            "excluded_tests": len(excluded_rows),
            "responses_per_participant": 60,
            "included_prolific_ids_unique": len({row["PROLIFIC_PID"] for row in included_rows}),
        }
        print(
            f"{study_id}: included={len(included_rows)}, excluded_tests={len(excluded_rows)}, "
            f"workbook={workbook_path}"
        )

    (args.output_root / "audit_export_summary.json").write_text(
        json.dumps(export_summary, indent=2), encoding="utf-8"
    )
    (args.output_root / "SENSITIVE_DATA_NOTICE.txt").write_text(
        "These workbooks contain Prolific participant and session identifiers.\n"
        "Keep them private and do not commit them to a public repository.\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
